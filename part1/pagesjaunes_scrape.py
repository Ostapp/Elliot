#!/usr/bin/env python
# -*- coding: utf-8 -*-
import requests
import os
# from pyvirtualdisplay import Display
#from bs4 import BeautifulSoup
from openpyxl import load_workbook
from user_agent import generate_user_agent
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import json
import random
import time
import subprocess
import pickle
import re
import sys
import xlsxwriter
import tldextract
import argparse

# shub deploy 240421
# sudo docker run -p 5023:5023 -p 8050:8050 -p 8051:8051 scrapinghub/splash --max-timeout 120

'''
I decided to use OOP approach to solve this task the most efficiently. 
The Page object is responsible for general rendering of any pages using random user agent headers, closing pop up windows and restarting external VPN script if the bot is blocked by the website. However, the VPN restart feature turned out to be not needed in the end, since the website failed to recoginze my bot as a bot.
The SearchPage object inherits from the Page object and represents a page with search results. It stores pharmacies links, names and locates link to the next search page. 
The InfoPage inherits from the Page object and represents a pharmacie profile. By scrolling it to the bottom it makes sure that the profile page renders fully. It also stores a website's URL if available. 
The Spider object is responsible for the most efficient click execution via ActionChains which turned out to be necessary for processing SearchPage object.
The Administrator object puts everything together. It loads and stores the list of the cities to be processed, creates search URLs, prepares proxies (if needed) and random user agents to be used by the Page object. It launches the search, renders the pages, executes the crawling itself by guiding the Spider object. It also performs backup saving of the crawling process so that it could be restarted from the last point in case of accident crashes. Finally, it saves the results in pickle and .xlsx formats.
In the end I designed a simple terminal interface with the help of argparse module. The interface allows to point to specific files containing list of the cities to be processed, it also allows to choose a phrase to be searched for as well as the name of a final file containing crawling results and if the proximite option is to be activated during the search or not. Finally, the most important feature of the interface is that it allows to restart the script from the backup file, so that the progress is not lost in case of accident crashes. That is quite important since crawling of this website turned out to be a quite time consuming process.

'''

class Administrator():

	def __init__(self, search_phrase = "pharmacie", last_start_url_index = 0, proximite = "0" , cities_filename="France 50 biggest cities.xlsx", output_filename="results"):
		self.cities = self.get_cities(cities_filename)
		# self.proxies = self.open_proxies_list()
		# self.proxy_gen = self.get_proxy()
		self.headers = self.headers_ua()
		self.start_urls = self.start_search_urls(search_phrase,proximite)
		self.links = {}
		self.spider = Spider()
		self.last_start_url_index = last_start_url_index
		self.search_phrase = search_phrase
		self.proximite = proximite
		self.cities_filename = cities_filename
		self.output_filename = output_filename

	def get_cities(self, cities_filename):

		cities = []

		wb = load_workbook(filename = cities_filename)
		sheet = wb.get_active_sheet()

		col = 1
		for row in range(1,sheet.max_row+1):
			cities.append(sheet.cell(column = col, row = row).value)

		return cities

	def open_proxies_list(self):
		with open('thebigproxylist-17-09-03.txt','r') as f:
			proxies = f.readlines()

		stripped_proxies = []

		for proxy in proxies:
			proxy = proxy.strip()
			stripped_proxies.append(proxy)

		return stripped_proxies

	def random_user_agent(self):
		with open('user-agents.txt','r') as f:
			user_agents = f.readlines()
		user_agents = [h.rstrip('\n') for h in user_agents]
		random_index = random.randint(0,len(user_agents)-1)
		ua = user_agents[random_index]
		return ua

	def headers_ua(self):

		headers = requests.utils.default_headers()

		headers.update({
			'User-Agent': self.random_user_agent()
			})
		return headers

	def start_search_urls(self, search_phrase, proximite):

		urls = []

		for city in self.cities:
			url = 'https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui='+search_phrase+'&ou='+city+'&proximite='+proximite+'&quoiQuiInterprete='+search_phrase+'&carte=0'

			urls.append(url)

		return urls

	def process_start_url(self,url):

		'''
		It turned out that the website provides results randomly every time the search request is made. For example the first page search page for pharmacie in Paris, will always give almost very different list pharmacies in Paris. It also turned out that each of the next pagination pages had to be rendered as a separate Selenium webdriver instance. This also resulted in the fact that when I opened a profile page in the same tab and then returned to the previous search results page the search results page was not the same as previously, since it contained different results. The solution to open the profile pages in a new tab did not work either for some reason. To solve this problem I decided to do the following. For each of the search results pagination pages I opened 2 Selenium web driver instances. The first instance remained as a reference point to stay passive (master page) while the second page (work page) was used to access one profile link at once. Once the profile link was processed the work was closed and a new search page (work page) instance was opened to process the second profile page and so on until 20 profile pages were processed from on pagination search page. After that I extracted the next page url from the master page, closed the master page and rendered the next pagination page as a new master page with the correspondent work pages rendered one by one, until the next 20 results were processed. The idea was that if I process 20 pages from each of the pagination search pages I would eventually exhaust all of the search reults without omitting anything. It indeed worked like that. 
		'''

		print "processing " + url
		master_page = SearchPage(url)
		num_of_entries = len(master_page.names)
		temp_saved_records = {}
		temp_already_saved = 0

		while len(temp_saved_records) < num_of_entries:
			work_page = SearchPage(url)	
			for index in range(num_of_entries):
				random.shuffle(work_page.details_links) #shuffling the profiles for the browser to imitate chaotic scrolling by a user
				spider.center_on_button(work_page.browser, work_page.details_links[index])
				time.sleep(random.randint(1,5))

				name = work_page.names[index].text

				if name not in self.links:
					print name + " NAME IS UNIQUE"

					while work_page.browser.current_url == master_page.browser.current_url: #sometimes it was not enough to click once to open a profile page (info page), so I ordered the webdriver to click as many times as needed to open the info page
						try:
							spider.click(work_page.browser, work_page.details_links[index])
						except Exception as e:
							work_page.close_pop_ups(work_page.browser)
							try:
								spider.click(work_page.browser, work_page.details_links[index])
							except:
								pass
							pass

					time.sleep(random.randint(1,5)) # wait a bit for the info page to render

					info_page = InfoPage(work_page.browser)
					try:
						self.links[name] = info_page.link_to_website
						temp_saved_records[name] = info_page.link_to_website
						print str(len(temp_saved_records))
						print "LINK "+info_page.link_to_website+" APPENDED"
					except Exception as e:
						print e
						pass
					info_page.browser.close()
					time.sleep(1)
					work_page = SearchPage(url)

				elif temp_already_saved >= 20:
					print "THE PAGE HAS BEEN EXHAUSTED"
					work_page.browser.close()
					break
				else:
					print "*"*10 + "ENTRY ALREADY EXISTS" + "*"*10
					temp_already_saved+=1
					continue
			else:
				break
			try:
				work_page.browser.close()
			except Exception as e:
				print e
				pass

		if master_page.next_page:
			spider.center_on_button(master_page.browser,master_page.next_page)
			prev_url = master_page.browser.current_url
			while prev_url == master_page.browser.current_url: # here I make sure that the webdriver manages to go the next search pagination page successfully
				try:
					master_page.close_pop_ups(master_page.browser)
					spider.click(master_page.browser, master_page.next_page)
				except Exception as e:
					print e
					spider.click(master_page.browser, master_page.next_page)
			time.sleep(5)
			url = re.sub('contexte.*&','',master_page.browser.current_url)
			print "next page url " + url
			master_page.browser.close()
			try:
				work_page.browser.close()
			except:
				pass
			self.process_start_url(url)
		else:
			"*"*10 + "no next page" + "*"*10
			master_page.browser.close()
			self.save()
			try:
				work_page.browser.close()
			except:
				pass
			return

	def start(self):
		for url in self.start_urls[self.last_start_url_index:]:
			self.last_start_url_index = self.start_urls.index(url)
			self.process_start_url(url)
			self.backup_search_save()

	def save(self):
		'''saves results of the crawling'''

		with open (self.output_filename+'.pickle', 'w') as f:
			pickle.dump(self.links, f)
			print "***RESULTS SAVED***"

	def load(self):
		'''loads previously saved results of the crawling'''

		with open (self.output_filename+'.pickle','r') as f:
			results = pickle.load(f)
			return results
	
	def make_xlsx_file(self, output_filename):
		'''makes final xlsx file with the list of the links. This file is to be checked by the client for him to confirm if this part of the job is being performed in accordance with his expectations'''

		workbook = xlsxwriter.Workbook(self.output_filename+'.xlsx')
		worksheet = workbook.add_worksheet()

		results = {value for value in self.links.values()}

		row = 0
		for result in results:
			worksheet.write_string(row, 0, result)
			row +=1
		workbook.close()

	# def save_unique_results(self):
	# 	results_to_process = self.load()
	# 	results = {value for value in results_to_process.values()}

	def backup_search_save(self):

		backup = {
		"last_start_url_index": admin.last_start_url_index,
		"search_phrase" : admin.search_phrase,
		"proximite" : admin.proximite,
		"cities_filename" : admin.cities_filename,
		"output_filename" : admin.output_filename,
		"links" : admin.links
		}
		with open ('backup.pickle','w') as f:
			pickle.dump(backup, f)
			print "***BACKUP SAVED***"

class Page(object):

	def __init__(self, url=None, render=True):
		if render == True:
			self.browser = self.render_page(url)
		self.url = url

	def close_pop_ups(self, browser):
		try:
			browser.find_element_by_xpath('/html/body/div[1]/div/button/span').click()
		except:
			pass
		try:
			browser.find_element_by_xpath('/html/body/div[1]/div/button').click()
		except:
			pass

		try:
			browser.find_element_by_id('acc-alert-close').click()
		except:
			pass
		try:
			close = browser.find_elements_by_id('kamClose')
			for button in close:
				button.click()
		except:
			pass
		try:
			browser.find_element_by_class_name('pjpopin-closer-grandePopin').click()
		except:
			pass
		try:
			browser.find_element_by_class_name('pjpopin-closer').click()
		except:
			pass
		try:
			browser.find_element_by_class_name('kclose').click()
		except:
			pass
		try: 
			browser.find_element_by_class_name('lien-fermer').click()
		except:
			pass

	def render_page(self, url):

		def run_proxy_browser(url):

			def restart_vpn():
				'''if the bot is blocked by the website, the webdriver instance should be closed, the external VPN is to be restarted and the webdriver instance relaunched'''

				try:
					browser.close()
				except:
					pass
				try:
					os.system('''ps axf | grep hma-vpn.sh | grep -v grep | awk '{print "kill " $1 }' | sh''')
					time.sleep(10)
				except: 
					pass
				try:
					subprocess.Popen(["exec gnome-terminal -e 'bash -c \"sudo bash hma-vpn.sh -c id-file; exec bash\"'"], stdout=subprocess.PIPE, shell=True)
					time.sleep(10)
				except:
					return restart_vpn()

			# PROXY = next(admin.proxy_gen)
			# print PROXY
			profile = webdriver.FirefoxProfile()
			profile.set_preference("general.useragent.override",admin.random_user_agent())
			profile.set_preference("http.response.timeout", 30)
			profile.set_preference("dom.max_script_run_time", 30) # makes sure that the DOM tree does not load eternally.
			webdriver.DesiredCapabilities.FIREFOX['marionette'] = False
			# webdriver.DesiredCapabilities.FIREFOX['proxy']={
			# 	"httpProxy":PROXY,
			# 	"ftpProxy":PROXY,
			# 	"sslProxy":PROXY,
			# 	"noProxy":None,
			# 	"proxyType":"MANUAL",
			# 	"autodetect":False
			# }

			browser = webdriver.Firefox(profile)
			browser.set_page_load_timeout(120)

			try:
				browser.get(url)
			except Exception as e:
				print e
				browser.close()
				return run_proxy_browser(url)

			if browser.title == u'Problem loading page':
				print "Problem loading page"
				print browser.page_source
				browser.close()
				return run_proxy_browser(url)
			elif browser.title == u'You have been blocked':
				print u'You have been blocked'
				restart_vpn()
				return run_proxy_browser(url)
			elif u"JE SUIS UN HUMAIN" in browser.page_source:
				print "JE SUIS UN HUMAIN"
				restart_vpn()
				return run_proxy_browser(url)
			elif u'PAGESJAUNES.FR protège son contenu contre les robots et le réserve aux êtres humains.' in browser.page_source:
				print "JE SUIS UN HUMAIN"
				restart_vpn()
				return run_proxy_browser(url)
			try:
				if len(browser.page_source) < 100:
					browser.close()
					restart_vpn()
					return run_proxy_browser(url)
				else:
					pass
			except:
				return run_proxy_browser(url)

			self.close_pop_ups(browser)
			browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
			time.sleep(3)

			return browser

		time.sleep(random.randint(1,5))
		browser = run_proxy_browser(url)

		return browser

class SearchPage(Page):

	def __init__(self, url):
		super(SearchPage, self).__init__(url,render=True)
		try:
			self.details_links = self.browser.find_elements_by_class_name('details-links ')
			self.names = self.browser.find_elements_by_class_name('denomination-links')
		except Exception as e:
			'''re-render the page if the inital rendering failed'''
			print e
			self.browser.close()
			self.browser = self.render_page(url)
		try:
			self.next_page = self.browser.find_element_by_id('pagination-next')
		except Exception as e:
			print e
			print "*"*10 + "no next page during rendering page detected" + "*"*10
			self.next_page = 0
		
class InfoPage(Page):
	
	def __init__(self, browser):
		super(InfoPage, self).__init__(render=False)
		self.browser = browser
		self.browser.set_page_load_timeout(60)
		self.height = random.randint(50,400)
		self.browser.execute_script("window.scrollTo(0, "+ str(self.height) + ");")
		try:
			self.link_to_website = str(self.browser.find_element_by_class_name('bloc-info-sites-reseaux').find_element_by_class_name('value').text)
		except:
			self.link_to_website = "no link to website available"

class Spider():

	def center_on_button(self, browser, element):

		location_x = element.location['x']
		location_y = element.location['y'] - 200
		
		browser.execute_script("window.scrollTo(0, %d);" %location_x)
		browser.execute_script("window.scrollTo(0, %d);" %location_y)

	def click(self, browser, element):
		ActionChains(browser).click(element).perform()

	def center_on_button_and_click(self, browser, element):
		ActionChains(browser).move_to_element(element).click().perform()

if __name__ == "__main__":

	dir_path = os.path.dirname(os.path.realpath(__file__))
	backup_file_path = dir_path + '/backup.pickle'
	backup_exists = os.path.exists(backup_file_path)

	parser = argparse.ArgumentParser(description='scrape pagesjaunes')
	parser.add_argument('--search_phrase', default = "pharmacie")
	parser.add_argument('--cities_filename', default = "France 50 biggest cities.xlsx")
	parser.add_argument('--output_filename', default = "results")
	parser.add_argument('--p', action='store_const', const="1", default="0")
	parser.add_argument('--start_from_backup', action = 'store_const', const = 'yes', default="no")
	args = parser.parse_args()

	if args.start_from_backup == 'yes' and backup_exists == True:
		print "***LOADING FROM BACKUP***"
		try:
			with open ("backup.pickle", 'r') as f:
				backup = pickle.load(f)
		except Exception as e:
			print e
		search_phrase = backup['search_phrase']
		proximite = backup['proximite']
		cities_filename = backup['cities_filename']
		output_filename = backup['output_filename']
		last_start_url_index = backup['last_start_url_index']

		admin = Administrator(search_phrase=search_phrase, last_start_url_index=last_start_url_index, proximite=proximite, cities_filename=cities_filename, output_filename=output_filename)
		admin.links = backup['links']
	elif args.start_from_backup == 'yes' and backup_exists == True:
		print "***NO BACKUP FILE FOUND***"
		sys.exit()
	else:
		search_phrase = args.search_phrase
		proximite = args.p
		cities_filename = args.cities_filename
		output_filename = args.output_filename
		admin = Administrator(search_phrase = search_phrase, proximite = proximite, cities_filename = cities_filename, output_filename= output_filename)

	# subprocess.Popen(["exec gnome-terminal -e 'bash -c \"sudo bash hma-vpn.sh -c id-file; exec bash\"'"], stdout=subprocess.PIPE, shell=True)
	spider = Spider()
	# with Display(visible=0, size=(1024, 768)):
	admin.start()